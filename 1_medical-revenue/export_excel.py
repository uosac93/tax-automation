"""
병의원 매출 회계처리 내보내기 모듈
분개장 엑셀 + 더존 Smart A import용 CSV 생성
"""
import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from journal_engine import JournalEntry, format_won


# ============================================================
# 스타일 상수
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
SUBTITLE_FONT = Font(name="맑은 고딕", size=11, bold=True)
BODY_FONT = Font(name="맑은 고딕", size=10)
NUMBER_FORMAT = '#,##0'
NEGATIVE_FONT = Font(name="맑은 고딕", size=10, color="FF0000")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

# 매출 유형별 색상
TYPE_FILLS = {
    '요양급여-매출': PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    '의료급여-매출': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    '요양급여-본인부담차감': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    '의료급여-본인부담차감': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    '요양급여-미수금': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    '의료급여-미수금': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    '요양급여-입금': PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    '의료급여-입금': PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    '요양급여-원천징수': PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"),
    '의료급여-원천징수': PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"),
    '요양급여-미수금회수': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    '의료급여-미수금회수': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
}


def generate_journal_excel(entries, summary, output_path, institution=''):
    """분개장 엑셀 파일 생성"""
    wb = Workbook()

    # ── 시트1: 분개장 ──
    ws1 = wb.active
    ws1.title = "분개장"

    ws1.merge_cells('A1:J1')
    ws1['A1'] = f"분개장 — {institution}" if institution else "분개장"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER

    ws1.merge_cells('A2:J2')
    ws1['A2'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1['A2'].font = Font(name="맑은 고딕", size=9, color="888888")
    ws1['A2'].alignment = CENTER

    # 헤더
    headers = ['No', '전표일자', '적요', '차변계정', '차변금액', '대변계정', '대변금액',
               '매출유형', '거래처', '귀속월']
    col_widths = [6, 14, 32, 18, 16, 18, 16, 18, 12, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터
    for i, entry in enumerate(entries, 1):
        row = i + 4
        dr_label = f"{entry.debit_account} {entry.debit_name}".strip() if entry.debit_account else ''
        cr_label = f"{entry.credit_account} {entry.credit_name}".strip() if entry.credit_account else ''

        values = [
            i,
            entry.date,
            entry.description,
            dr_label,
            entry.debit_amount if entry.debit_amount != 0 else '',
            cr_label,
            entry.credit_amount if entry.credit_amount != 0 else '',
            entry.revenue_type,
            entry.partner_name,
            entry.month
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

            if col_idx in (5, 7):  # 금액 열
                if isinstance(val, (int, float)) and val != '':
                    cell.number_format = NUMBER_FORMAT
                    if val < 0:
                        cell.font = NEGATIVE_FONT
                cell.alignment = RIGHT
            elif col_idx == 1:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

            # 매출 유형별 색상
            type_fill = TYPE_FILLS.get(entry.revenue_type)
            if type_fill and col_idx == 8:
                cell.fill = type_fill

    # 합계 행
    total_row = len(entries) + 5
    ws1.cell(row=total_row, column=3, value="합 계").font = SUBTITLE_FONT
    ws1.cell(row=total_row, column=3).border = THIN_BORDER
    ws1.cell(row=total_row, column=3).alignment = CENTER

    total_debit = sum(e.debit_amount for e in entries)
    total_credit = sum(e.credit_amount for e in entries)

    cell_dr = ws1.cell(row=total_row, column=5, value=total_debit)
    cell_dr.font = SUBTITLE_FONT
    cell_dr.number_format = NUMBER_FORMAT
    cell_dr.alignment = RIGHT
    cell_dr.border = THIN_BORDER

    cell_cr = ws1.cell(row=total_row, column=7, value=total_credit)
    cell_cr.font = SUBTITLE_FONT
    cell_cr.number_format = NUMBER_FORMAT
    cell_cr.alignment = RIGHT
    cell_cr.border = THIN_BORDER

    # ── 시트2: 매출 요약 ──
    ws2 = wb.create_sheet("매출요약")

    ws2.merge_cells('A1:H1')
    ws2['A1'] = f"월별 매출 요약 — {institution}" if institution else "월별 매출 요약"
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = CENTER

    sum_headers = ['진료월', '요양급여수입', '의료급여수입', '본인부담차감',
                   '미수금발생', '입금액', '원천징수', '매출합계']
    sum_widths = [14, 18, 18, 16, 16, 16, 14, 18]

    for col_idx, (header, width) in enumerate(zip(sum_headers, sum_widths), 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    for i, (month, data) in enumerate(summary.items(), 1):
        row = i + 3
        values = [
            month,
            data.get('insurance', 0),
            data.get('medical_aid', 0),
            data.get('general_deduct', 0),
            data.get('ar_amount', 0),
            data.get('deposit', 0),
            data.get('tax_total', 0),
            data.get('total', 0)
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_idx >= 2:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = RIGHT
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = NEGATIVE_FONT
            else:
                cell.alignment = CENTER

    # 총합계
    total_row2 = len(summary) + 4
    ws2.cell(row=total_row2, column=1, value="합 계").font = SUBTITLE_FONT
    ws2.cell(row=total_row2, column=1).alignment = CENTER
    ws2.cell(row=total_row2, column=1).border = THIN_BORDER

    sum_keys = ['insurance', 'medical_aid', 'general_deduct', 'ar_amount',
                'deposit', 'tax_total', 'total']
    for col_idx, key in enumerate(sum_keys, 2):
        total_val = sum(s.get(key, 0) for s in summary.values())
        cell = ws2.cell(row=total_row2, column=col_idx, value=total_val)
        cell.font = SUBTITLE_FONT
        cell.number_format = NUMBER_FORMAT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER

    wb.save(output_path)
    return output_path


def generate_douzone_csv(entries, output_path):
    """
    더존 Smart A import용 CSV 생성
    형식: 전표일자, 구분, 계정코드, 계정명, 금액, 거래처코드, 거래처명, 적요
    """
    with open(output_path, 'w', newline='', encoding='cp949') as f:
        writer = csv.writer(f)
        writer.writerow(['전표일자', '구분', '계정코드', '계정명', '금액',
                        '거래처코드', '거래처명', '적요'])

        for entry in entries:
            date_str = entry.date.replace('-', '')

            # 차변
            if entry.debit_amount != 0:
                writer.writerow([
                    date_str,
                    '차변',
                    entry.debit_account,
                    entry.debit_name,
                    entry.debit_amount,
                    entry.partner_code,
                    entry.partner_name,
                    entry.description
                ])
            # 대변
            if entry.credit_amount != 0:
                writer.writerow([
                    date_str,
                    '대변',
                    entry.credit_account,
                    entry.credit_name,
                    entry.credit_amount,
                    entry.partner_code,
                    entry.partner_name,
                    entry.description
                ])

    return output_path


def generate_wehago_xls(cash_records, output_path):
    """
    위하고 매입매출전표 업로드용 .xls 생성
    병의원 매출: 카드(18-카면), 현금영수증(23-현면), 현금(20-면건)
    """
    import xlwt
    import calendar

    wb = xlwt.Workbook(encoding='cp949')
    ws = wb.add_sheet('매출자료 & 매입자료')

    # 헤더
    headers = [
        '년도', '월', '일', '매입매출구분(1-매출/2-매입)',
        '과세유형', '불공제사유', '신용카드거래처코드', '신용카드사명',
        '신용카드(가맹점)번호', '거래처명', '사업자(주민)등록번호',
        '공급가액', '부가세', '품명', '전자세금(1.전자)',
        '기본계정', '상대계정', '현금영수증 승인번호'
    ]

    header_style = xlwt.easyxf(
        'font: bold on; align: horiz center; '
        'borders: left thin, right thin, top thin, bottom thin; '
        'pattern: pattern solid, fore_colour light_blue;'
    )
    body_style = xlwt.easyxf(
        'align: horiz center; '
        'borders: left thin, right thin, top thin, bottom thin;'
    )
    num_style = xlwt.easyxf(
        'align: horiz right; '
        'borders: left thin, right thin, top thin, bottom thin;',
        num_format_str='#,##0'
    )

    for c, h in enumerate(headers):
        ws.write(0, c, h, header_style)
        ws.col(c).width = 256 * max(len(h) + 2, 12)

    row = 1
    for rec in cash_records:
        month_str = rec.get('month', '')
        if not month_str or len(month_str) < 7:
            continue

        year = int(month_str[:4])
        month = int(month_str[5:7])
        last_day = calendar.monthrange(year, month)[1]

        card_amt = rec.get('card_amount', 0)
        receipt_amt = rec.get('receipt_amount', 0)
        cash_amt = rec.get('cash_amount', 0)

        # 카드매출 → 18 (카면)
        if card_amt > 0:
            ws.write(row, 0, year, body_style)
            ws.write(row, 1, month, body_style)
            ws.write(row, 2, last_day, body_style)
            ws.write(row, 3, 1, body_style)  # 매출
            ws.write(row, 4, 18, body_style)  # 카면
            ws.write(row, 5, '', body_style)
            ws.write(row, 6, '', body_style)
            ws.write(row, 7, '', body_style)
            ws.write(row, 8, '', body_style)
            ws.write(row, 9, '', body_style)
            ws.write(row, 10, '', body_style)
            ws.write(row, 11, card_amt, num_style)
            ws.write(row, 12, 0, num_style)
            ws.write(row, 13, '카드매출', body_style)
            ws.write(row, 14, '', body_style)
            ws.write(row, 15, '', body_style)
            ws.write(row, 16, '', body_style)
            ws.write(row, 17, '', body_style)
            row += 1

        # 현금영수증 → 23 (현면)
        if receipt_amt > 0:
            ws.write(row, 0, year, body_style)
            ws.write(row, 1, month, body_style)
            ws.write(row, 2, last_day, body_style)
            ws.write(row, 3, 1, body_style)
            ws.write(row, 4, 23, body_style)  # 현면
            ws.write(row, 5, '', body_style)
            ws.write(row, 6, '', body_style)
            ws.write(row, 7, '', body_style)
            ws.write(row, 8, '', body_style)
            ws.write(row, 9, '', body_style)
            ws.write(row, 10, '', body_style)
            ws.write(row, 11, receipt_amt, num_style)
            ws.write(row, 12, 0, num_style)
            ws.write(row, 13, '현금영수증매출', body_style)
            ws.write(row, 14, '', body_style)
            ws.write(row, 15, '', body_style)
            ws.write(row, 16, '', body_style)
            ws.write(row, 17, '', body_style)
            row += 1

        # 현금매출 → 20 (면건)
        if cash_amt > 0:
            ws.write(row, 0, year, body_style)
            ws.write(row, 1, month, body_style)
            ws.write(row, 2, last_day, body_style)
            ws.write(row, 3, 1, body_style)
            ws.write(row, 4, 20, body_style)  # 면건
            ws.write(row, 5, '', body_style)
            ws.write(row, 6, '', body_style)
            ws.write(row, 7, '', body_style)
            ws.write(row, 8, '', body_style)
            ws.write(row, 9, '', body_style)
            ws.write(row, 10, '', body_style)
            ws.write(row, 11, cash_amt, num_style)
            ws.write(row, 12, 0, num_style)
            ws.write(row, 13, '현금매출', body_style)
            ws.write(row, 14, '', body_style)
            ws.write(row, 15, '', body_style)
            ws.write(row, 16, '', body_style)
            ws.write(row, 17, '', body_style)
            row += 1

    wb.save(output_path)
    return output_path


def generate_all(entries, summary, output_dir, institution=''):
    """엑셀 + CSV 모두 생성"""
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime('%Y%m%d')
    inst_name = institution or '병의원'

    excel_path = os.path.join(output_dir, f"분개장_{inst_name}_{date_str}.xlsx")
    csv_path = os.path.join(output_dir, f"더존import_{inst_name}_{date_str}.csv")

    generate_journal_excel(entries, summary, excel_path, institution)
    generate_douzone_csv(entries, csv_path)

    return {'excel': excel_path, 'csv': csv_path}
