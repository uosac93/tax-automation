"""
법인세 검토 결과 리포트 생성 모듈
콘솔 출력 + Excel 파일 생성
"""
import os
import sys
import io
from datetime import datetime

# Windows 콘솔 인코딩 문제 해결
if sys.platform == 'win32' and sys.stdout and hasattr(sys.stdout, 'buffer') and sys.stdout.buffer:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tax_reviewer import ReviewItem, format_won


# 스타일 상수
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
ISSUE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NORMAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CHECK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
SUBTITLE_FONT = Font(name="맑은 고딕", size=11, bold=True)
BODY_FONT = Font(name="맑은 고딕", size=10)
NUMBER_FORMAT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def print_console_report(data, results):
    """콘솔에 검토 결과 출력"""
    info = data.get("회사정보", {})
    tax = data.get("세액조정", {})

    print("\n" + "=" * 80)
    print("  법인세 신고서 검토 결과 보고서")
    print("=" * 80)

    # 회사 기본정보
    print(f"\n■ 회사정보")
    print(f"  법인명:       {info.get('법인명', 'N/A')}")
    print(f"  사업자등록번호: {info.get('사업자등록번호', 'N/A')}")
    print(f"  대표자:       {info.get('대표자', 'N/A')}")
    print(f"  사업연도:     {info.get('사업연도_시작', '')} ~ {info.get('사업연도_종료', '')}")
    print(f"  업태/종목:    {info.get('업태', '')} / {info.get('종목', '')}")

    # 핵심 세무 수치
    print(f"\n■ 핵심 세무 수치")
    print(f"  결산서상 당기순손익:  {format_won(tax.get('결산서상당기순손익'))}원")
    print(f"  익금산입:            {format_won(tax.get('익금산입'))}원")
    print(f"  손금산입:            {format_won(tax.get('손금산입', 0))}원")
    print(f"  각사업연도소득금액:  {format_won(tax.get('각사업연도소득금액'))}원")
    print(f"  과세표준:            {format_won(tax.get('과세표준'))}원")
    print(f"  산출세액:            {format_won(tax.get('산출세액'))}원 (세율 {tax.get('세율', 'N/A')}%)")
    print(f"  공제감면세액:        {format_won(tax.get('최저한세적용대상_공제감면세액'))}원")
    print(f"  차감세액:            {format_won(tax.get('차감세액'))}원")
    print(f"  기납부세액:          {format_won(tax.get('기납부세액'))}원")
    print(f"  차감납부할세액:      {format_won(tax.get('차감납부할세액'))}원")
    print(f"  분납세액:            {format_won(tax.get('분납세액'))}원")

    # 농어촌특별세
    rural = data.get("농어촌특별세", {})
    if rural.get("산출세액"):
        print(f"\n■ 농어촌특별세")
        print(f"  과세표준: {format_won(rural.get('과세표준'))}원")
        print(f"  산출세액: {format_won(rural.get('산출세액'))}원")

    # 검토 결과 요약
    이슈수 = sum(1 for r in results if r.상태 == "이슈")
    정상수 = sum(1 for r in results if r.상태 == "정상")
    확인필요수 = sum(1 for r in results if r.상태 == "확인필요")

    print(f"\n■ 검토 결과 요약")
    print(f"  총 검토항목: {len(results)}개")
    print(f"  [O] 정상:     {정상수}개")
    print(f"  [!] 확인필요: {확인필요수}개")
    print(f"  [X] 이슈:     {이슈수}개")

    # 이슈/확인필요 항목 먼저 표시
    이슈_items = [r for r in results if r.상태 == "이슈"]
    확인필요_items = [r for r in results if r.상태 == "확인필요"]

    # ANSI 색상 코드
    RED = "\033[91m"
    YELLOW = "\033[93m"
    GREEN = "\033[92m"
    RESET = "\033[0m"

    def print_item(item):
        if item.상태 == "이슈":
            color = RED
            icon = "X"
        elif item.상태 == "확인필요":
            color = YELLOW
            icon = "!"
        elif item.상태 == "정상":
            color = GREEN
            icon = "O"
        else:
            color = ""
            icon = "-"
        print(f"  {color}[{icon}] {item.항목명}{RESET}")
        # 비고에서 [서식명] 부분과 금액/설명 부분 분리
        if item.비고:
            비고 = item.비고
            # [] 서식명 추출
            import re
            서식들 = re.findall(r'\[([^\]]+)\]', 비고)
            if 서식들:
                print(f"    서식: {' ↔ '.join(서식들)}")
            # 금액 비교 표시
            if item.신고서금액 is not None and item.검증금액 is not None:
                print(f"    신고서: {format_won(item.신고서금액)}")
                print(f"    검  증: {format_won(item.검증금액)}")
                if item.차이금액 and item.차이금액 > 0:
                    print(f"    차  이: {format_won(item.차이금액)}원")
            elif item.신고서금액 is not None:
                print(f"    금  액: {format_won(item.신고서금액)}")
            # [] 제거 후 나머지 설명만 표시
            설명 = re.sub(r'\[[^\]]+\]\s*', '', 비고).strip()
            # 금액 패턴도 제거 (이미 위에서 표시)
            if 설명 and len(설명) > 5:
                print(f"    내  용: {설명}")
        elif item.신고서금액 is not None:
            print(f"    금  액: {format_won(item.신고서금액)}")
        print()

    if 이슈_items:
        print(f"\n{'=' * 60}")
        print(f"  {RED}[X] 이슈 항목{RESET}")
        print(f"{'=' * 60}")
        for item in 이슈_items:
            print_item(item)

    if 확인필요_items:
        print(f"\n{'=' * 60}")
        print(f"  {YELLOW}[!] 확인필요 항목{RESET}")
        print(f"{'=' * 60}")
        for item in 확인필요_items:
            print_item(item)

    # 정상 항목
    정상_items = [r for r in results if r.상태 == "정상"]
    if 정상_items:
        print(f"\n{'=' * 60}")
        print(f"  {GREEN}[O] 정상 항목{RESET}")
        print(f"{'=' * 60}")
        for item in 정상_items:
            print_item(item)

    # 총평
    print(f"\n{'=' * 80}")
    print("  총평")
    print(f"{'=' * 80}")
    if 이슈수 == 0:
        print("  주요 세액 계산 항목에서 이슈가 발견되지 않았습니다.")
        print("  세무조정계산서의 산술적 정합성은 양호합니다.")
    else:
        print(f"  {이슈수}건의 이슈가 발견되었습니다. 상세 내용을 확인하시기 바랍니다.")

    if 확인필요수 > 0:
        print(f"  {확인필요수}건의 수동 확인이 필요한 항목이 있습니다.")

    print("=" * 80)


def generate_excel_report(data, results, output_path):
    """Excel 검토 보고서 생성 - 상세검토결과 시트 1개"""
    wb = Workbook()
    ws = wb.active
    ws.title = "상세검토결과"

    info = data.get("회사정보", {})
    nm = info.get("법인명", "")
    period = f"{info.get('사업연도_시작', '')}~{info.get('사업연도_종료', '')}"

    # 헤더
    headers = ["카테고리", "검토항목", "상태", "신고서금액", "검증금액", "차이금액", "비고"]
    header_font = Font(name="맑은 고딕", size=10, bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 데이터
    for i, item in enumerate(results, 2):
        ws.cell(row=i, column=1, value=item.카테고리).font = BODY_FONT
        ws.cell(row=i, column=2, value=item.항목명).font = BODY_FONT
        ws.cell(row=i, column=3, value=item.상태).font = BODY_FONT

        if item.신고서금액 is not None:
            ws.cell(row=i, column=4, value=item.신고서금액).number_format = NUMBER_FORMAT
        if item.검증금액 is not None:
            ws.cell(row=i, column=5, value=item.검증금액).number_format = NUMBER_FORMAT
        if item.차이금액 is not None and item.차이금액 > 0:
            ws.cell(row=i, column=6, value=item.차이금액).number_format = NUMBER_FORMAT

        비고_text = (item.비고 or "").replace("\n", " | ")
        ws.cell(row=i, column=7, value=비고_text).font = BODY_FONT

        for col in range(1, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER

    # 열 너비
    col_widths = [14, 35, 10, 18, 18, 18, 80]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 저장
    wb.save(output_path)
    print(f"\n  Excel 보고서 저장 완료: {output_path}")
